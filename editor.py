import tkinter as tk
from tkinter import filedialog, messagebox
import json
from openpyxl import load_workbook

database = []

def get_data():
    q = entry_q.get()
    answers = []
    for i in range(8):
        t = ans[i].get()
        p = pts[i].get()
        if t:
            try: p = int(p)
            except: p = 0
            answers.append({"text": t.upper(), "points": p})
    return q, answers

def preview():
    canvas.delete("all")
    q, a = get_data()

    canvas.create_text(300,20,text=q,fill="white")
    for i in range(8):
        x = 50 if i<4 else 350
        y = 60 + (i%4)*60
        canvas.create_rectangle(x,y,x+250,y+50,fill="yellow")
        if i<len(a):
            canvas.create_text(x+10,y+15,text=a[i]["text"],anchor="w")
            canvas.create_text(x+220,y+15,text=str(a[i]["points"]))
        else:
            canvas.create_text(x+120,y+15,text=str(i+1))

def add():
    q,a = get_data()
    if not q or not a:
        messagebox.showwarning("Error","Data kosong")
        return
    database.append({"question":q,"answers":a})
    listbox.insert(tk.END,q)
    clear()

def clear():
    entry_q.delete(0,tk.END)
    for e in ans: e.delete(0,tk.END)
    for p in pts: p.delete(0,tk.END)
    preview()

def save():
    file = filedialog.asksaveasfilename(defaultextension=".json")
    if file:
        with open(file,"w") as f:
            json.dump(database,f,indent=2)

def load():
    global database
    file = filedialog.askopenfilename(filetypes=[("JSON","*.json")])
    if file:
        with open(file) as f:
            database = json.load(f)
        listbox.delete(0,tk.END)
        for d in database:
            listbox.insert(tk.END,d["question"])

def import_excel():
    file = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
    if not file: return
    wb = load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        q = row[0]
        answers = []
        for i in range(1,len(row),2):
            if row[i]:
                answers.append({
                    "text": str(row[i]).upper(),
                    "points": int(row[i+1] or 0)
                })
        if q and answers:
            database.append({"question":q,"answers":answers})
            listbox.insert(tk.END,q)

root = tk.Tk()
root.title("Editor Family Feud")

tk.Label(root,text="Pertanyaan").grid(row=0,column=0)
entry_q = tk.Entry(root,width=50)
entry_q.grid(row=0,column=1,columnspan=4)

ans=[]; pts=[]
for i in range(8):
    tk.Label(root,text=f"Ans {i+1}").grid(row=i+1,column=0)
    e=tk.Entry(root); e.grid(row=i+1,column=1)
    p=tk.Entry(root,width=5); p.grid(row=i+1,column=2)
    e.bind("<KeyRelease>",lambda e:preview())
    p.bind("<KeyRelease>",lambda e:preview())
    ans.append(e); pts.append(p)

tk.Button(root,text="Tambah",command=add).grid(row=10,column=0)
tk.Button(root,text="Clear",command=clear).grid(row=10,column=1)
tk.Button(root,text="Import Excel",command=import_excel).grid(row=10,column=2)
tk.Button(root,text="Save",command=save).grid(row=10,column=3)
tk.Button(root,text="Load",command=load).grid(row=10,column=4)

listbox = tk.Listbox(root,width=40)
listbox.grid(row=1,column=3,rowspan=8)

canvas = tk.Canvas(root,width=600,height=350,bg="blue")
canvas.grid(row=11,column=0,columnspan=5)

preview()
root.mainloop()
